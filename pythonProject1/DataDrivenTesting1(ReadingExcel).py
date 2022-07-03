#Data Driven Testing
#Openpyxl ile Excel tablolarını okumak
#ReadingExcel

"""""
1-) Excel'den veri oku

2-) Excel'e veri nasıl yazılır

3-) Veri odaklı test senaryosu

"""

import openpyxl
file = r"C:\Users\oguzh\Documents\Data.xlsx"
workbook = openpyxl.load_workbook(file)

sheet = workbook["Sheet1"]
rows = sheet.max_row
columns = sheet.max_column

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='       ')
    print()